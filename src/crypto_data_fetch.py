import requests
import pandas as pd
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Fetch top 50 cryptocurrencies from CoinMarketCap API
def fetch_crypto_data(api_key):
    url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest"
    headers = {
        'Accepts': 'application/json',
        'X-CMC_PRO_API_KEY': api_key,
    }
    params = {
        'start': '1',
        'limit': '50',
        'convert': 'USD'
    }
    
    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()  # Raise an error for bad responses
        return response.json()['data']  # Get the data field from the response
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return []

# Perform data analysis
def analyze_data(data):
    # Convert to DataFrame for easier analysis
    df = pd.DataFrame(data)
    
    # Extract relevant fields from the 'quote' column
    df['current_price'] = df['quote'].apply(lambda x: x['USD']['price'])
    df['market_cap'] = df['quote'].apply(lambda x: x['USD']['market_cap'])
    df['total_volume'] = df['quote'].apply(lambda x: x['USD']['volume_24h'])
    df['percent_change_24h'] = df['quote'].apply(lambda x: x['USD']['percent_change_24h'])
    
    # Identify top 5 cryptocurrencies by market cap
    top_5_by_market_cap = df.nlargest(5, 'market_cap')[['name', 'market_cap']]
    
    # Calculate the average price of the top 50 cryptocurrencies
    average_price = df['current_price'].mean()
    
    # Find the highest and lowest 24-hour percentage price change
    highest_24h_change = df.nlargest(1, 'percent_change_24h')[['name', 'percent_change_24h']]
    lowest_24h_change = df.nsmallest(1, 'percent_change_24h')[['name', 'percent_change_24h']]
    
    return {
        'top_5_by_market_cap': top_5_by_market_cap,
        'average_price': average_price,
        'highest_24h_change': highest_24h_change,
        'lowest_24h_change': lowest_24h_change
    }

# Update Excel with live data
def update_excel(data, file_name="crypto_data.xlsx"):
    # Convert data to DataFrame
    df = pd.DataFrame(data)
    
    # Extract relevant fields from the 'quote' column
    df['current_price'] = df['quote'].apply(lambda x: x['USD']['price'])
    df['market_cap'] = df['quote'].apply(lambda x: x['USD']['market_cap'])
    df['total_volume'] = df['quote'].apply(lambda x: x['USD']['volume_24h'])
    df['percent_change_24h'] = df['quote'].apply(lambda x: x['USD']['percent_change_24h'])
    
    # Load or create Excel workbook
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
    
    # Clear the existing sheet and add headers
    sheet.delete_rows(1, sheet.max_row)
    headers = ['Name', 'Symbol', 'Current Price (USD)', 'Market Cap', '24h Trading Volume', 'Price Change (24h %)']
    sheet.append(headers)
    
    # Add new rows of data
    for row in dataframe_to_rows(df[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'percent_change_24h']], index=False, header=False):
        sheet.append(row)
    
    # Save Excel file
    workbook.save(file_name)

# Main function to continuously update data and Excel every 5 minutes
def main():
    api_key = "YOUR_API_KEY"  # Replace with your CoinMarketCap API key
    while True:
        print("Fetching cryptocurrency data...")
        data = fetch_crypto_data('60097829-ca9c-43f3-8c5c-791852fed1d7')
        
        # Perform analysis
        analysis = analyze_data(data)
        
        if analysis['top_5_by_market_cap'] is not None:
            print("\nAnalysis Report:")
            print(f"Top 5 by Market Cap:\n{analysis['top_5_by_market_cap']}")
            print(f"Average Price: ${analysis['average_price']:.2f}")
            print(f"Highest 24h Change:\n{analysis['highest_24h_change']}")
            print(f"Lowest 24h Change:\n{analysis['lowest_24h_change']}")
        else:
            print("No data available for analysis.")
        
        # Update Excel sheet
        update_excel(data)
        
        print("Excel sheet updated. Waiting for 5 minutes before the next update...")
        time.sleep(60)  # Wait for 1 minute

if __name__ == "__main__":
    main()
